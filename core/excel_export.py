from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_to_excel(items, keyword: str, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d_%H%M")
    safe_keyword = keyword.replace("/", "_").replace(" ", "_")
    out_path = output_dir / f"{safe_keyword}_{today}.xlsx"

    rows = []
    for rank, it in enumerate(items, start=1):
        rows.append({
            "順位": rank,
            "サイト": it.site,
            "タイトル": it.title,
            "価格": it.price,
            "状態": it.condition or "",
            "発送元": it.location or "",
            "画像URL": it.image_url or "",
            "商品リンク": it.item_url,
        })
    df = pd.DataFrame(rows)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ranking", index=False)
        ws = writer.sheets["ranking"]

        header_fill = PatternFill("solid", fgColor="305496")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, _ in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        widths = {"順位": 6, "サイト": 16, "タイトル": 60, "価格": 12, "状態": 12, "発送元": 20, "画像URL": 40, "商品リンク": 60}
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 20)

        ws.freeze_panes = "A2"

        link_col = list(df.columns).index("商品リンク") + 1
        for r in range(2, len(rows) + 2):
            link_cell = ws.cell(row=r, column=link_col)
            url = link_cell.value
            if url:
                link_cell.hyperlink = url
                link_cell.font = Font(color="0563C1", underline="single")

    return out_path
